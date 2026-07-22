"""通用资产导入脚本 — 智能识别多种格式和非标准列名

设计目标:
  - 客户手上的资产数据格式千差万别 (Excel/JSON/非标准CSV/纯文本IP列表)
  - 优先复用标准导入流程 (import_assets + load_csv)
  - 当标准导入返回 0 条记录时, 自动降级到通用解析

三段降级策略:
  Step 1: 标准导入 — 直接调 import_assets + load_csv 试加载
  Step 2: 列名预检 — 读表头判断标准列名是否全部缺失
  Step 3: 通用解析 — 智能映射列名 → 生成标准 CSV → 再调 import_assets

支持格式:
  - CSV (.csv)
  - Excel (.xlsx, 用 openpyxl, 不依赖 pandas)
  - JSON (.json, list of dict 或 dict of dict)
  - 纯文本 (.txt, 每行一个 IP 或 "IP,hostname")

CLI 用法:
  # 自动识别格式和 layer
  python3 import_assets_flexible.py /path/to/assets.xlsx

  # 指定 layer
  python3 import_assets_flexible.py /path/to/assets.csv --layer tenant

  # 预览模式 (只看识别结果, 不导入)
  python3 import_assets_flexible.py /path/to/assets.json --dry-run

  # 纯文本 IP 列表
  python3 import_assets_flexible.py /path/to/iplist.txt
"""
from __future__ import annotations
import csv
import io
import json
import os
import re
import sys
from pathlib import Path
from typing import Any

# 同目录导入
_SCRIPT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(_SCRIPT_DIR))
from asset_resolver import (  # noqa: E402
    AssetResolver,
    import_assets,
)


# ============================================================
# 智能列名映射表
# ============================================================
# 每个标准字段对应一组关键词, 列名(归一化后)包含任一关键词即匹配
# 优先级: 列表中靠前的关键词优先级更高 (用于消歧)
FIELD_KEYWORDS: dict[str, list[str]] = {
    # 内网 IP (主键) — 优先匹配 "内网/private", 避免误匹配公网
    "ip": ["内网ip", "内网地址", "privateip", "lanip", "私网ip",
           "ip地址", "ip", "主机ip", "内网"],
    # 公网 IP — 必须含 "公网/public/外网/wan"
    "public_ip": ["公网ip", "公网地址", "publicip", "wanip", "外网ip",
                  "外网地址", "公网"],
    "hostname": ["主机名", "hostname", "hostname", "实例名", "名称",
                 "instancename", "servername", "name"],
    # AppID (租户层关键字段)
    "appid": ["appid", "appid", "app_id", "租户id", "租户appid", "账号id"],
    # VPC/网络
    "network": ["网络", "vpc", "network", "vpcid", "网络名"],
    # 实例 ID (ins-xxx)
    "instance_id": ["实例id", "instanceid", "instance_id", "ins-"],
    # 资产 ID
    "asset_id": ["uuid", "主机id", "hostid", "host_id", "资产id", "机器id"],
    # OS
    "os": ["操作系统", "os", "镜像", "image", "imagename", "镜像名称"],
    "zone": ["可用区", "zone", "区域"],
    "cpu": ["cpu", "核数", "核"],
    "memory": ["内存", "memory", "ram"],
    "owner": ["创建者", "owner", "归属", "账号id"],
    "status": ["状态", "status", "运行状态"],
    "host_ip": ["宿主机", "宿主机ip", "hostip", "物理机ip"],
}


# 标准 CSV 列名 (与 asset_resolver._parse_row 一致)
STANDARD_COLUMNS_PLATFORM = [
    "主机ID", "主机名", "IP地址", "宿主机内网IP", "可用区",
    "操作系统名称", "镜像ID", "CPU", "内存(GB)",
    "系统盘（类型#大小GB）", "数据盘（类型#大小GB）",
    "创建者账号ID", "状态", "创建时间",
]

STANDARD_COLUMNS_TENANT = [
    "UUID", "主机名", "实例ID", "创建者账号ID", "AppID", "状态",
    "内网地址IP", "公网IP地址", "IPv6地址", "宿主机内网IP",
    "可用区", "网络", "镜像名称", "镜像ID",
    "CPU（核）", "内存（GB）",
    "系统盘（类型#大小GB）", "数据盘（类型#大小GB）", "创建时间",
]


# ============================================================
# 文件格式识别
# ============================================================
def _read_file(source: Path) -> tuple[list[str], list[dict]]:
    """读取文件, 返回 (表头列表, 行字典列表)

    支持: CSV / Excel / JSON / 纯文本

    对于纯文本 IP 列表, 构造最小表头 ["ip"]
    """
    suffix = source.suffix.lower()

    if suffix == ".csv":
        return _read_csv(source)
    if suffix in (".xlsx", ".xlsm"):
        return _read_excel(source)
    if suffix == ".json":
        return _read_json(source)
    if suffix in (".txt", ".text", ".log") or _looks_like_text(source):
        return _read_text_iplist(source)

    # 默认按 CSV 尝试
    print(f"[WARN] 未识别的文件后缀 '{suffix}', 按 CSV 尝试", file=sys.stderr)
    return _read_csv(source)


def _read_csv(source: Path) -> tuple[list[str], list[dict]]:
    with open(source, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames or []
        rows = list(reader)
    return list(headers), rows


def _read_excel(source: Path) -> tuple[list[str], list[dict]]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError(
            "读取 Excel 需要安装 openpyxl: pip install openpyxl"
        ) from e

    wb = load_workbook(source, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = [str(h) if h is not None else "" for h in next(rows_iter)]
    except StopIteration:
        return [], []
    data = []
    for row in rows_iter:
        if all(c is None for c in row):
            continue
        data.append({headers[i]: ("" if row[i] is None else str(row[i]))
                     for i in range(min(len(headers), len(row)))})
    wb.close()
    return headers, data


def _read_json(source: Path) -> tuple[list[str], list[dict]]:
    with open(source, "r", encoding="utf-8") as f:
        data = json.load(f)
    if isinstance(data, dict):
        # dict of dict: 以 key 为标识, value 为字段
        if all(isinstance(v, dict) for v in data.values()):
            rows = []
            for key, fields in data.items():
                row = dict(fields)
                row.setdefault("name", key)
                row.setdefault("ip", key)
                rows.append(row)
            headers = list(rows[0].keys()) if rows else []
            return headers, rows
        data = [data]
    if not isinstance(data, list):
        raise ValueError(f"JSON 格式不支持: {type(data)}")
    rows = [dict(item) for item in data if isinstance(item, dict)]
    headers = list(rows[0].keys()) if rows else []
    return headers, rows


def _looks_like_text(source: Path) -> bool:
    """检测是否纯文本 IP 列表 (每行一个 IP 或空行)"""
    try:
        with open(source, "r", encoding="utf-8") as f:
            for i, line in enumerate(f):
                if i > 50:
                    break
                line = line.strip()
                if not line:
                    continue
                # 不是 IP 或 IP,hostname 格式 → 不是纯文本 IP 列表
                if not re.match(r"^\d{1,3}(\.\d{1,3}){3}", line):
                    return False
            return True
    except Exception:
        return False


def _read_text_iplist(source: Path) -> tuple[list[str], list[dict]]:
    headers = ["ip"]
    rows = []
    with open(source, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            # 支持 "IP,hostname" 或 "IP hostname" 格式
            parts = re.split(r"[,\s]+", line, maxsplit=1)
            row = {"ip": parts[0].strip()}
            if len(parts) > 1 and parts[1].strip():
                row["hostname"] = parts[1].strip()
            rows.append(row)
    if rows and "hostname" in rows[0]:
        headers = ["ip", "hostname"]
    return headers, rows


# ============================================================
# 智能列名映射
# ============================================================
def _normalize(name: str) -> str:
    """列名归一化: 去空格、小写、去括号"""
    if not name:
        return ""
    s = str(name).strip().lower().replace(" ", "")
    s = re.sub(r"[（(].*?[)）]", "", s)  # 去括号内容
    return s


def build_column_mapping(headers: list[str]) -> dict[str, str]:
    """构建 客户列名 → 标准字段 的映射

    Returns:
        { 客户原始列名: 标准字段名 }
    """
    mapping: dict[str, str] = {}
    used_fields: set[str] = set()

    # 优先匹配高优先级字段 (ip 优先于 public_ip)
    for field in ["ip", "public_ip", "host_ip", "hostname", "appid",
                  "instance_id", "asset_id", "network", "os", "zone",
                  "cpu", "memory", "owner", "status"]:
        keywords = FIELD_KEYWORDS.get(field, [])
        for header in headers:
            if header in mapping:
                continue
            norm = _normalize(header)
            if not norm:
                continue
            for kw in keywords:
                if kw in norm:
                    mapping[header] = field
                    used_fields.add(field)
                    break
            if field in used_fields:
                break

    return mapping


def infer_layer(mapping: dict[str, str], headers: list[str]) -> str:
    """根据匹配到的字段推断 layer"""
    fields = set(mapping.values())
    # 有 appid → tenant
    if "appid" in fields:
        return "tenant"
    # 有 host_ip / 宿主机 → platform
    if "host_ip" in fields:
        return "platform"
    # 有 instance_id (ins-xxx) → tenant
    if "instance_id" in fields:
        return "tenant"
    # 默认 tenant
    return "tenant"


# ============================================================
# 生成标准 CSV
# ============================================================
def _get_value(row: dict, client_col: str, mapping: dict[str, str]) -> str:
    """从行数据取值, 支持映射后的列名"""
    if client_col in row:
        return str(row[client_col] or "").strip()
    return ""


def generate_standard_csv(
    rows: list[dict],
    mapping: dict[str, str],
    layer: str,
    output_path: Path,
) -> int:
    """按标准列名生成 CSV, 供 asset_resolver.load_csv 读取

    Returns:
        写入的记录数
    """
    if layer == "platform":
        standard_cols = STANDARD_COLUMNS_PLATFORM
    else:
        standard_cols = STANDARD_COLUMNS_TENANT

    # 反转映射: 标准字段 → 客户列名
    field_to_client: dict[str, str] = {v: k for k, v in mapping.items()}

    n = 0
    with open(output_path, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=standard_cols)
        writer.writeheader()
        for row in rows:
            out: dict[str, str] = {col: "" for col in standard_cols}
            # 映射字段
            for client_col, field in mapping.items():
                value = _get_value(row, client_col, mapping)
                if not value:
                    continue
                # 标准字段 → 标准列名
                if layer == "platform":
                    col = _field_to_platform_col(field, value)
                else:
                    col = _field_to_tenant_col(field, value)
                if col and col in out:
                    out[col] = value
            # 至少要有 ip 或 hostname 或 asset_id 才写入
            has_key = any(out.get(k) for k in
                          ["IP地址", "内网地址IP", "主机名", "主机ID", "UUID"])
            if has_key:
                writer.writerow(out)
                n += 1
    return n


def _field_to_platform_col(field: str, value: str) -> str:
    """标准字段 → platform CSV 列名"""
    m = {
        "asset_id": "主机ID",
        "hostname": "主机名",
        "ip": "IP地址",
        "host_ip": "宿主机内网IP",
        "zone": "可用区",
        "os": "操作系统名称",
        "cpu": "CPU",
        "memory": "内存(GB)",
        "owner": "创建者账号ID",
        "status": "状态",
    }
    return m.get(field, "")


def _field_to_tenant_col(field: str, value: str) -> str:
    """标准字段 → tenant CSV 列名"""
    m = {
        "asset_id": "UUID",
        "hostname": "主机名",
        "instance_id": "实例ID",
        "ip": "内网地址IP",
        "public_ip": "公网IP地址",
        "host_ip": "宿主机内网IP",
        "zone": "可用区",
        "network": "网络",
        "os": "镜像名称",
        "cpu": "CPU（核）",
        "memory": "内存（GB）",
        "owner": "创建者账号ID",
        "appid": "AppID",
        "status": "状态",
    }
    return m.get(field, "")


# ============================================================
# 三段降级主入口
# ============================================================
def smart_import(source_file: str | Path, layer: str = "auto") -> dict:
    """智能资产导入 (三段降级)

    Args:
        source_file: 资产文件路径 (CSV/Excel/JSON/纯文本)
        layer: "auto" / "platform" / "tenant"

    Returns:
        {
            "method": "standard" | "flexible",
            "layer": "platform" | "tenant",
            "total": int,
            "mapped_columns": dict,   # flexible 模式下的列名映射
            "target_file": str,       # 导入后的目标文件路径
            "warnings": list,
        }
    """
    source = Path(source_file).expanduser().resolve()
    if not source.exists():
        raise FileNotFoundError(f"资产文件不存在: {source}")

    warnings: list[str] = []
    tmp_csv: Path | None = None

    # 非 CSV 格式 → 直接走通用解析
    if source.suffix.lower() not in (".csv",):
        # 检查是否标准格式失败 → 直接通用解析
        result = _do_flexible_import(source, layer, warnings)
        return result

    # CSV 格式 → 先尝试标准导入
    print(f"[Step 1] 尝试标准导入: {source}", file=sys.stderr)
    try:
        # 临时确定 layer (auto 时读表头判断)
        headers, _ = _read_csv(source)
        if layer == "auto":
            # 检查是否含标准列名
            has_platform = any(h in STANDARD_COLUMNS_PLATFORM for h in headers)
            has_tenant = any(h in STANDARD_COLUMNS_TENANT for h in headers)
            if "AppID" in headers or "内网地址IP" in headers:
                layer = "tenant"
            elif "主机ID" in headers or "IP地址" in headers:
                layer = "platform"
            else:
                layer = "tenant"  # 默认
            print(f"[Step 1] 推断 layer={layer}", file=sys.stderr)

        target = import_assets(source, layer=layer)

        # 用 AssetResolver 试加载, 检查有效记录数
        resolver = AssetResolver()
        n = resolver.load_csv(target, layer=layer)
        print(f"[Step 1] 标准导入结果: {n} 条有效记录", file=sys.stderr)

        if n > 0:
            return {
                "method": "standard",
                "layer": layer,
                "total": n,
                "mapped_columns": {},
                "target_file": str(target),
                "warnings": [],
            }
        # n == 0 → 列名不匹配, 降级
        warnings.append(
            f"标准导入返回 0 条记录, 可能列名不匹配 (表头: {headers})"
        )
    except Exception as e:
        warnings.append(f"标准导入失败: {e}")

    # 降级到通用解析
    print("[Step 3] 降级到通用解析", file=sys.stderr)
    return _do_flexible_import(source, layer, warnings)


def _do_flexible_import(
    source: Path, layer: str, warnings: list[str]
) -> dict:
    """执行通用解析导入"""
    print(f"[Step 3] 读取文件: {source}", file=sys.stderr)
    headers, rows = _read_file(source)
    if not rows:
        raise ValueError(f"文件中无有效数据行: {source}")

    print(f"[Step 3] 识别到表头: {headers}", file=sys.stderr)
    print(f"[Step 3] 识别到 {len(rows)} 行数据", file=sys.stderr)

    # 智能列名映射
    mapping = build_column_mapping(headers)
    print(f"[Step 3] 列名映射: {mapping}", file=sys.stderr)

    if not mapping:
        raise ValueError(
            f"无法识别任何列名, 请检查文件格式. 表头: {headers}"
        )

    # 检查关键字段
    if "ip" not in mapping.values() and "hostname" not in mapping.values():
        # 纯文本 IP 列表时 headers=["ip"], mapping 应含 ip
        if "ip" not in mapping.values():
            raise ValueError(
                "未识别到 IP 或主机名列, 无法导入. "
                f"识别到的字段: {list(mapping.values())}"
            )

    # 推断 layer
    if layer == "auto":
        layer = infer_layer(mapping, headers)
        print(f"[Step 3] 推断 layer={layer}", file=sys.stderr)

    # 生成标准 CSV 到临时文件
    tmp_csv = source.parent / f"_converted_{layer}_{source.stem}.csv"
    n = generate_standard_csv(rows, mapping, layer, tmp_csv)
    print(f"[Step 3] 生成标准 CSV: {tmp_csv} ({n} 条)", file=sys.stderr)

    if n == 0:
        raise ValueError("转换后 0 条记录, 请检查数据是否有效")

    # 调标准导入
    target = import_assets(tmp_csv, layer=layer)

    # 验证导入
    resolver = AssetResolver()
    loaded = resolver.load_csv(target, layer=layer)
    print(f"[Step 3] 导入验证: {loaded} 条有效记录", file=sys.stderr)

    # 清理临时文件
    try:
        tmp_csv.unlink()
    except OSError:
        pass

    return {
        "method": "flexible",
        "layer": layer,
        "total": loaded,
        "mapped_columns": mapping,
        "target_file": str(target),
        "warnings": warnings,
    }


# ============================================================
# CLI
# ============================================================
def main():
    import argparse

    parser = argparse.ArgumentParser(
        description="通用资产导入脚本 — 智能识别多种格式和非标准列名",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  # 自动识别格式和 layer
  python3 import_assets_flexible.py /path/to/assets.xlsx

  # 指定 layer
  python3 import_assets_flexible.py /path/to/assets.csv --layer tenant

  # 预览模式 (只看识别结果, 不导入)
  python3 import_assets_flexible.py /path/to/assets.json --dry-run

  # 纯文本 IP 列表
  python3 import_assets_flexible.py /path/to/iplist.txt
        """,
    )
    parser.add_argument("source", help="资产文件路径 (CSV/Excel/JSON/纯文本)")
    parser.add_argument(
        "--layer",
        choices=["auto", "platform", "tenant"],
        default="auto",
        help="资产层 (默认 auto, 自动推断)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="只预览识别结果, 不实际导入",
    )

    args = parser.parse_args()

    if args.dry_run:
        return _dry_run(args.source, args.layer)

    # 检查环境变量
    if not os.environ.get("CODEBUDDY_PLUGIN_DATA"):
        print("[ERROR] CODEBUDDY_PLUGIN_DATA 环境变量未设置, 无法导入",
              file=sys.stderr)
        print("[INFO] 如需预览, 请加 --dry-run 参数", file=sys.stderr)
        sys.exit(1)

    try:
        result = smart_import(args.source, layer=args.layer)
        print("\n" + "=" * 60, file=sys.stderr)
        print(f"导入完成", file=sys.stderr)
        print(f"  方式: {result['method']}", file=sys.stderr)
        print(f"  层级: {result['layer']}", file=sys.stderr)
        print(f"  记录数: {result['total']}", file=sys.stderr)
        print(f"  目标文件: {result['target_file']}", file=sys.stderr)
        if result["mapped_columns"]:
            print(f"  列名映射:", file=sys.stderr)
            for client_col, field in result["mapped_columns"].items():
                print(f"    {client_col} → {field}", file=sys.stderr)
        if result["warnings"]:
            print(f"  警告:", file=sys.stderr)
            for w in result["warnings"]:
                print(f"    - {w}", file=sys.stderr)
        print("=" * 60, file=sys.stderr)
    except Exception as e:
        print(f"[ERROR] 导入失败: {e}", file=sys.stderr)
        sys.exit(1)


def _dry_run(source_path: str, layer: str):
    """预览模式: 只识别不导入"""
    source = Path(source_path).expanduser().resolve()
    if not source.exists():
        print(f"[ERROR] 文件不存在: {source}", file=sys.stderr)
        sys.exit(1)

    print(f"=== 预览模式 (不导入) ===", file=sys.stderr)
    print(f"文件: {source}", file=sys.stderr)
    print(f"格式: {source.suffix}", file=sys.stderr)

    headers, rows = _read_file(source)
    print(f"\n表头 ({len(headers)} 列):", file=sys.stderr)
    for h in headers:
        print(f"  - {h}", file=sys.stderr)

    print(f"\n数据行数: {len(rows)}", file=sys.stderr)
    if rows:
        print(f"\n首行样例:", file=sys.stderr)
        for k, v in list(rows[0].items())[:8]:
            print(f"  {k}: {v}", file=sys.stderr)

    mapping = build_column_mapping(headers)
    print(f"\n智能列名映射:", file=sys.stderr)
    if mapping:
        for client_col, field in mapping.items():
            print(f"  {client_col} → {field}", file=sys.stderr)
    else:
        print("  (未识别到任何可映射列名)", file=sys.stderr)

    inferred = layer if layer != "auto" else infer_layer(mapping, headers)
    print(f"\n推断 layer: {inferred}", file=sys.stderr)

    # 检查关键字段
    fields = set(mapping.values())
    if "ip" in fields or "hostname" in fields:
        print(f"\n[OK] 包含关键字段 ({', '.join(sorted(fields & {'ip', 'hostname'}))})", file=sys.stderr)
    else:
        print(f"\n[WARN] 未识别到 IP 或主机名列", file=sys.stderr)

    print(f"\n=== 预览完成, 如需导入请去掉 --dry-run ===", file=sys.stderr)


if __name__ == "__main__":
    main()
